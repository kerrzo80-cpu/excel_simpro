#!/usr/bin/env python3
"""Remove accidental external workbook references from formulas.

After copying the Works area from an archived workbook, Excel may keep formulas
pointing back to the donor workbook, e.g.:
  'https://.../2117 - garry smith.xlsm'!tblTaskLibrary[Task Code]

This script removes those external prefixes so formulas point to local tables:
  tblTaskLibrary[Task Code]

Usage:
  python3 fix_external_workbook_links.py --yes
"""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Quoting/Works Pricing Tool V3 BACKUP before visual style 2026-05-27 09-11.xlsm"

# Matches quoted external workbook references ending in .xlsm'! before table refs.
EXTERNAL_PREFIX_RE = re.compile(r"'[^']+\.xlsm'!")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(WORKBOOK))
    parser.add_argument("--yes", action="store_true")
    args = parser.parse_args()

    workbook = Path(args.workbook).expanduser()
    if not workbook.exists():
        print("Workbook not found:", workbook)
        return 1

    print("Workbook:", workbook)
    print("This will remove external .xlsm workbook prefixes from formulas only.")
    if not args.yes:
        answer = input("Continue? [y/N]: ").strip().lower()
        if answer not in {"y", "yes"}:
            print("Cancelled.")
            return 0

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup = workbook.with_name(workbook.stem + f" BACKUP before external link fix {stamp}" + workbook.suffix)
    shutil.copy2(workbook, backup)
    print("Backup created:", backup)

    wb = load_workbook(workbook, keep_vba=True)
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and ".xlsm'!" in value:
                    new_value = EXTERNAL_PREFIX_RE.sub("", value)
                    if new_value != value:
                        cell.value = new_value
                        changed += 1

    wb.save(workbook)
    print("Formulas fixed:", changed)
    print("Workbook saved.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
