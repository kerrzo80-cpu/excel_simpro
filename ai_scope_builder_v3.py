#!/usr/bin/env python3
"""
AI Scope Builder v3 - workbook-aware.

What this does now:
  - Reads the real Task Library from the Works Pricing Tool workbook.
  - Combines rule-based scope building with fuzzy matching against real task names/codes.
  - Shows which suggested task codes actually exist in the workbook.
  - Runs without editing the workbook by default.

Safe first tests:
  python3 ai_scope_builder_v3.py "remove basin and cap off supplies"
  python3 ai_scope_builder_v3.py "replace radiator with new double panel radiator"
  python3 ai_scope_builder_v3.py --workbook "/path/to/Works Pricing Tool.xlsm" "install outside tap"

Next stage will be --insert, but this version is deliberately read-only.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

import ai_scope_builder_v2 as rules

DEFAULT_WORKBOOKS = [
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Quoting/Works Pricing Tool V3 BACKUP before visual style 2026-05-27 09-11.xlsm",
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Pricing/Works Pricing Tool.xlsm",
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Quoting/Works Pricing Tool V2.xlsm",
]

TASK_SHEET = "Task Library"
FIRST_TASK_ROW = 6

ACTION_WORDS = {"remove", "take", "rip", "strip", "disconnect", "move", "relocate", "shift", "replace", "swap", "change", "renew", "install", "fit", "cap", "blank", "unblock", "repair"}
GENERIC_WORDS = {"work", "works", "job", "customer", "wants", "needs", "need", "please", "quote", "price", "new", "old", "existing", "supply", "supplies", "with", "and", "the", "a", "an", "to", "from", "in", "on", "for", "of"}
OBJECT_SYNONYMS = {
    "sink": "basin",
    "whb": "basin",
    "wc": "toilet",
    "loo": "toilet",
    "rad": "radiator",
    "rads": "radiator",
    "trv": "valve",
    "trvs": "valve",
    "pipes": "pipework",
    "pipe": "pipework",
}


@dataclass(frozen=True)
class WorkbookTask:
    row: int
    code: str
    category: str
    name: str
    driver_type: str
    driver_unit: str
    base_hours: float | int | str | None
    hours_per_unit: float | int | str | None
    minimum_hours: float | int | str | None
    notes: str
    tokens: list[str]


@dataclass(frozen=True)
class V3Suggestion:
    code: str
    name: str
    confidence: int
    source: str
    exists_in_workbook: bool
    workbook_name: str | None
    reason: str


@dataclass(frozen=True)
class V3Result:
    description: str
    workbook_path: str
    confidence: int
    suggestions: list[V3Suggestion]
    questions: list[str]
    warnings: list[str]


def clean(value: object) -> str:
    return str(value or "").strip()


def norm(text: object) -> str:
    text = clean(text).lower().replace("/", " ").replace("-", " ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokenise(text: object) -> list[str]:
    output: list[str] = []
    for token in norm(text).split():
        if token in GENERIC_WORDS:
            continue
        token = OBJECT_SYNONYMS.get(token, token)
        output.append(token)
    return output


def choose_workbook(cli_path: str | None = None) -> Path:
    if cli_path:
        path = Path(cli_path).expanduser()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path
    for path in DEFAULT_WORKBOOKS:
        if path.exists():
            return path
    raise FileNotFoundError("Could not find Works Pricing Tool workbook. Use --workbook /path/to/file.xlsm")


def read_task_library(workbook_path: Path) -> list[WorkbookTask]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True, keep_vba=True)
    if TASK_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {TASK_SHEET}")
    ws = wb[TASK_SHEET]
    tasks: list[WorkbookTask] = []
    blank_streak = 0

    for row_no in range(FIRST_TASK_ROW, ws.max_row + 1):
        values = [ws.cell(row=row_no, column=col).value for col in range(1, 10)]
        code = clean(values[0])
        name = clean(values[2])
        if not code and not name:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0
        if not code or not name:
            continue

        category = clean(values[1])
        driver_type = clean(values[3])
        driver_unit = clean(values[4])
        notes = clean(values[8])
        haystack = " ".join([code, category, name, driver_type, driver_unit, notes])
        tasks.append(WorkbookTask(
            row=row_no,
            code=code,
            category=category,
            name=name,
            driver_type=driver_type,
            driver_unit=driver_unit,
            base_hours=values[5],
            hours_per_unit=values[6],
            minimum_hours=values[7],
            notes=notes,
            tokens=tokenise(haystack),
        ))
    return tasks


def contains_phrase(text: str, phrase: str) -> bool:
    return re.search(rf"\b{re.escape(norm(phrase))}\b", norm(text)) is not None


def infer_intent(description: str) -> str:
    d = norm(description)
    if any(contains_phrase(d, p) for p in ["remove", "take out", "rip out", "strip out", "disconnect"]):
        return "remove"
    if any(contains_phrase(d, p) for p in ["move", "relocate", "shift"]):
        return "move"
    if any(contains_phrase(d, p) for p in ["replace", "swap", "change", "renew", "new"]):
        return "replace"
    if any(contains_phrase(d, p) for p in ["install", "fit", "supply and fit"]):
        return "install"
    if any(contains_phrase(d, p) for p in ["cap", "blank", "make safe"]):
        return "cap"
    if any(contains_phrase(d, p) for p in ["unblock", "blocked", "blockage"]):
        return "unblock"
    if any(contains_phrase(d, p) for p in ["repair", "fix", "not working", "not filling", "leaking"]):
        return "repair"
    return "unknown"


def score_workbook_task(description: str, task: WorkbookTask) -> int:
    desc_tokens = set(tokenise(description))
    task_tokens = set(task.tokens)
    if not desc_tokens:
        return 0

    score = 0
    overlap = desc_tokens & task_tokens
    score += len(overlap) * 22

    d_norm = norm(description)
    name_norm = norm(task.name)
    code_norm = norm(task.code)

    if name_norm and name_norm in d_norm:
        score += 100
    if d_norm and d_norm in name_norm:
        score += 70

    score += int(SequenceMatcher(None, d_norm, name_norm).ratio() * 25)
    score += int(SequenceMatcher(None, d_norm, code_norm).ratio() * 8)

    intent = infer_intent(description)
    if intent == "remove":
        if {"remove", "cap", "disconnect"} & task_tokens:
            score += 45
        if {"move", "install", "fit"} & task_tokens and "remove" not in task_tokens:
            score -= 40
    elif intent == "move":
        if {"move", "relocate", "alter"} & task_tokens:
            score += 45
        if "remove" in task_tokens:
            score -= 30
    elif intent == "replace":
        if {"replace", "fit", "install"} & task_tokens:
            score += 35
    elif intent == "install":
        if {"fit", "install"} & task_tokens:
            score += 45
    elif intent in {"repair", "unblock"}:
        if {"replace", "repair", "unblock", "valve", "flush", "fill"} & task_tokens:
            score += 35

    return max(0, score)


def task_by_code(tasks: Sequence[WorkbookTask]) -> dict[str, WorkbookTask]:
    return {t.code.upper(): t for t in tasks}


def merge_suggestions(description: str, tasks: Sequence[WorkbookTask]) -> tuple[list[V3Suggestion], list[str], list[str], int]:
    workbook_codes = task_by_code(tasks)
    rule_result = rules.base.build_scope(description)
    suggestions: dict[str, V3Suggestion] = {}
    warnings = list(rule_result.warnings)
    questions = list(rule_result.questions)

    # 1. Start with rule/scope suggestions, but verify against workbook.
    for item in rule_result.suggested_tasks:
        code = item.code.upper()
        wb_task = workbook_codes.get(code)
        exists = wb_task is not None
        suggestions[code] = V3Suggestion(
            code=code,
            name=wb_task.name if wb_task else item.name,
            confidence=item.confidence,
            source=item.source,
            exists_in_workbook=exists,
            workbook_name=wb_task.name if wb_task else None,
            reason=item.reason,
        )

    # 2. Add top real workbook matches.
    scored = sorted(((score_workbook_task(description, task), task) for task in tasks), key=lambda x: x[0], reverse=True)
    for score, task in scored[:8]:
        if score < 55:
            continue
        code = task.code.upper()
        confidence = min(92, max(45, score))
        existing = suggestions.get(code)
        if existing and existing.confidence >= confidence:
            continue
        suggestions[code] = V3Suggestion(
            code=code,
            name=task.name,
            confidence=confidence,
            source="workbook-match",
            exists_in_workbook=True,
            workbook_name=task.name,
            reason="Matched against the real Task Library in the workbook.",
        )

    # 3. Keep strongest suggestions but don't hide missing custom tasks.
    ordered = sorted(suggestions.values(), key=lambda s: (s.exists_in_workbook, s.confidence), reverse=True)

    missing = [s.code for s in ordered if not s.exists_in_workbook]
    if missing:
        msg = "Suggested task codes not currently in workbook Task Library: " + ", ".join(missing)
        if msg not in warnings:
            warnings.append(msg)

    if not ordered:
        confidence = 0
        warnings.append("No task suggestion found. Price this once, then save it as a learned rule.")
    else:
        confidence = round(sum(s.confidence for s in ordered[:5]) / min(5, len(ordered)))

    return ordered[:8], questions[:6], warnings, confidence


def build_scope_v3(description: str, workbook_path: Path) -> V3Result:
    tasks = read_task_library(workbook_path)
    suggestions, questions, warnings, confidence = merge_suggestions(description, tasks)
    return V3Result(
        description=description,
        workbook_path=str(workbook_path),
        confidence=confidence,
        suggestions=suggestions,
        questions=questions,
        warnings=warnings,
    )


def save_learned(description: str, codes: Sequence[str]) -> None:
    rules.base.save_learned_rule(description, codes)


def print_result(result: V3Result) -> None:
    print("")
    print(f"Description: {result.description}")
    print(f"Workbook: {result.workbook_path}")
    print(f"Overall confidence: {result.confidence}%")
    print("")

    if not result.suggestions:
        print("Suggested scope: none yet")
    else:
        print("Suggested scope:")
        for idx, item in enumerate(result.suggestions, 1):
            status = "in workbook" if item.exists_in_workbook else "missing from workbook"
            print(f"{idx}. {item.code} - {item.name} ({item.confidence}%, {status}, {item.source})")
            print(f"   Reason: {item.reason}")

    if result.questions:
        print("")
        print("Questions to confirm:")
        for question in result.questions:
            print(f"- {question}")

    if result.warnings:
        print("")
        print("Warnings:")
        for warning in result.warnings:
            print(f"- {warning}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("description", nargs="*", help="Plain-English description of works")
    parser.add_argument("--workbook", help="Path to Works Pricing Tool .xlsm")
    parser.add_argument("--json", action="store_true", help="Output JSON")
    parser.add_argument("--learn", action="store_true", help="Save learned mapping: --learn 'description' TASK1 TASK2")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.learn:
        if len(args.description) < 2:
            print("Usage: python3 ai_scope_builder_v3.py --learn 'remove basin' BASIN-REMOVE PIPE-CAP")
            return 2
        save_learned(args.description[0], args.description[1:])
        print(f"Learned rule saved for: {args.description[0]}")
        print("Task codes:", ", ".join(args.description[1:]))
        return 0

    description = " ".join(args.description).strip()
    if not description:
        description = input("Works description: ").strip()
    if not description:
        print("No description entered.")
        return 1

    workbook_path = choose_workbook(args.workbook)
    result = build_scope_v3(description, workbook_path)
    if args.json:
        print(json.dumps(asdict(result), indent=2))
    else:
        print_result(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
