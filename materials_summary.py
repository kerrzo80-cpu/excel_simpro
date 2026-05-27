#!/usr/bin/env python3
"""Materials Summary / Buying List.

Reads the Works Pricing Tool workbook and rolls task materials up into totals.
This solves the issue where each task sends little material bits like 0.5m pipe,
instead of one practical buying list such as 2 x 3m lengths.

Usage:
  python3 materials_summary.py
  python3 materials_summary.py --write
  python3 materials_summary.py --workbook "/path/to/Works Pricing Tool.xlsm" --write
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_WORKBOOKS = [
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Quoting/Works Pricing Tool V3 BACKUP before visual style 2026-05-27 09-11.xlsm",
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Pricing/Works Pricing Tool.xlsm",
    Path.home() / "Library/CloudStorage/OneDrive-ErrolWatsonGroup/Quoting/Works Pricing Tool V2.xlsm",
]

WORKS_SHEET = "Works To Be Carried Out"
TASK_MATERIALS_SHEET = "Task Materials"
MATERIALS_LIST_SHEET = "Materials List"
SUMMARY_SHEET = "Materials Summary"
BUYING_SHEET = "Buying List"
FIRST_WORK_ROW = 12
LAST_WORK_ROW = 250
PLACEHOLDER = "select work task"
PIPE_LENGTH_M = 3.0


@dataclass
class WorkLine:
    row: int
    task_name: str
    task_code: str
    count: float
    measurement: float


@dataclass
class MaterialTotal:
    code: str
    item: str
    spec: str
    unit: str
    qty: float
    sell_ex_vat: float
    source_tasks: set[str]


def clean(value: Any) -> str:
    return str(value or "").strip()


def clean_unit(value: Any, fallback: str = "") -> str:
    text = clean(value)
    if not text or text.startswith("=") or "XLOOKUP" in text.upper():
        text = clean(fallback)
    if not text or text.startswith("="):
        return "each"
    return text


def num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def choose_workbook(cli_path: str | None) -> Path:
    if cli_path:
        path = Path(cli_path).expanduser()
        if not path.exists():
            raise FileNotFoundError(path)
        return path
    for path in DEFAULT_WORKBOOKS:
        if path.exists():
            return path
    raise FileNotFoundError("Could not find workbook. Use --workbook /path/to/file.xlsm")


def sheet_by_stripped_name(wb, wanted: str):
    for ws in wb.worksheets:
        if ws.title.strip() == wanted.strip():
            return ws
    raise KeyError(f"Sheet not found: {wanted}")


def read_work_lines(wb) -> list[WorkLine]:
    ws = sheet_by_stripped_name(wb, WORKS_SHEET)
    lines: list[WorkLine] = []
    for row in range(FIRST_WORK_ROW, min(ws.max_row, LAST_WORK_ROW) + 1):
        task_name = clean(ws.cell(row, 3).value)
        task_code = clean(ws.cell(row, 4).value)
        if not task_name or not task_code:
            continue
        if PLACEHOLDER in task_name.lower() or PLACEHOLDER in task_code.lower():
            continue
        lines.append(
            WorkLine(
                row=row,
                task_name=task_name,
                task_code=task_code,
                count=max(num(ws.cell(row, 5).value, 1), 1),
                measurement=num(ws.cell(row, 6).value, 0),
            )
        )
    return lines


def read_material_prices(wb) -> dict[str, dict[str, Any]]:
    ws = sheet_by_stripped_name(wb, MATERIALS_LIST_SHEET)
    prices: dict[str, dict[str, Any]] = {}
    for row in range(11, ws.max_row + 1):
        code = clean(ws.cell(row, 1).value)
        if not code:
            continue
        prices[code] = {
            "item": clean(ws.cell(row, 4).value),
            "spec": clean(ws.cell(row, 5).value),
            "unit": clean(ws.cell(row, 6).value),
            "sell_ex_vat": num(ws.cell(row, 9).value, 0),
        }
    return prices


def read_task_materials(wb) -> dict[str, list[dict[str, Any]]]:
    ws = sheet_by_stripped_name(wb, TASK_MATERIALS_SHEET)
    task_materials: dict[str, list[dict[str, Any]]] = {}
    for row in range(6, ws.max_row + 1):
        task_code = clean(ws.cell(row, 1).value)
        material_code = clean(ws.cell(row, 2).value)
        if not task_code or not material_code:
            continue
        task_materials.setdefault(task_code, []).append(
            {
                "material_code": material_code,
                "material_item": clean(ws.cell(row, 3).value),
                "fixed_qty": num(ws.cell(row, 4).value, 0),
                "qty_per_unit": num(ws.cell(row, 5).value, 0),
                "unit": clean(ws.cell(row, 6).value),
                "unit_sell_ex_vat": num(ws.cell(row, 7).value, 0),
                "notes": clean(ws.cell(row, 8).value),
            }
        )
    return task_materials


def build_totals(wb) -> tuple[list[WorkLine], dict[str, MaterialTotal]]:
    work_lines = read_work_lines(wb)
    prices = read_material_prices(wb)
    task_materials = read_task_materials(wb)
    totals: dict[str, MaterialTotal] = {}

    for line in work_lines:
        for material in task_materials.get(line.task_code, []):
            code = material["material_code"]
            price = prices.get(code, {})
            qty = (material["fixed_qty"] + (material["qty_per_unit"] * line.measurement)) * line.count
            if qty <= 0:
                continue
            total = totals.get(code)
            if total is None:
                total = MaterialTotal(
                    code=code,
                    item=clean(price.get("item") or material["material_item"] or code),
                    spec=clean(price.get("spec")),
                    unit=clean_unit(price.get("unit"), material.get("unit")),
                    qty=0.0,
                    sell_ex_vat=num(material["unit_sell_ex_vat"] or price.get("sell_ex_vat"), 0),
                    source_tasks=set(),
                )
                totals[code] = total
            total.qty += qty
            total.source_tasks.add(line.task_code)

    return work_lines, totals


def is_pipe_or_length_material(total: MaterialTotal) -> bool:
    text = " ".join([total.code, total.item, total.spec, total.unit]).lower()
    return total.unit.lower() in {"m", "metre", "meter", "metres", "meters"} or "pipe" in text or "copper" in text


def buy_quantity(total: MaterialTotal) -> str:
    unit = total.unit.lower()
    qty = total.qty
    if is_pipe_or_length_material(total):
        lengths = max(1, math.ceil(qty / PIPE_LENGTH_M))
        return f"{lengths} x {PIPE_LENGTH_M:g}m length(s)"
    if unit in {"each", "ea", "unit", "kit", "roll", "tube", "pack", "pair", "bottle"}:
        return str(math.ceil(qty))
    return f"{qty:g} {total.unit}".strip()


def money(value: float) -> float:
    return round(float(value or 0), 2)


def print_summary(work_lines: list[WorkLine], totals: dict[str, MaterialTotal]) -> None:
    print(f"Work lines read: {len(work_lines)}")
    print(f"Unique materials: {len(totals)}")
    print("")
    print("Materials Summary")
    print("-----------------")
    for total in sorted(totals.values(), key=lambda t: (t.item.lower(), t.spec.lower(), t.code)):
        print(f"{total.code} | {total.item} {total.spec} | Qty: {total.qty:g} {total.unit} | Buy: {buy_quantity(total)}")


def ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
        return ws
    return wb.create_sheet(name)


def write_sheets(wb, totals: dict[str, MaterialTotal]) -> None:
    summary = ensure_sheet(wb, SUMMARY_SHEET)
    summary.append(["Material Code", "Item", "Spec", "Unit", "Total Required", "Buy Quantity", "Unit Sell Ex VAT", "Extended Sell Ex VAT", "Source Tasks"])

    buying = ensure_sheet(wb, BUYING_SHEET)
    buying.append(["Material Code", "Item", "Spec", "Buy Quantity", "Total Required", "Unit", "Notes"])

    for total in sorted(totals.values(), key=lambda t: (t.item.lower(), t.spec.lower(), t.code)):
        extended = money(total.qty * total.sell_ex_vat)
        sources = ", ".join(sorted(total.source_tasks))
        buy = buy_quantity(total)
        summary.append([total.code, total.item, total.spec, total.unit, round(total.qty, 3), buy, total.sell_ex_vat, extended, sources])
        buying.append([total.code, total.item, total.spec, buy, round(total.qty, 3), total.unit, ""])

    for ws in (summary, buying):
        ws.freeze_panes = "A2"
        widths = [18, 32, 28, 12, 16, 20, 16, 18, 40]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook")
    parser.add_argument("--write", action="store_true", help="Write Materials Summary and Buying List sheets into the workbook")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = choose_workbook(args.workbook)
    wb = load_workbook(workbook_path, keep_vba=True)
    work_lines, totals = build_totals(wb)
    print("Workbook:", workbook_path)
    print_summary(work_lines, totals)

    if args.write:
        write_sheets(wb, totals)
        wb.save(workbook_path)
        print("")
        print("Written sheets:", SUMMARY_SHEET, "and", BUYING_SHEET)
        print("Workbook saved.")
    else:
        print("")
        print("Read-only mode. Add --write to create/update workbook summary sheets.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
