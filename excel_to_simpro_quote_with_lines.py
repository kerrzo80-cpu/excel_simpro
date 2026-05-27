import getpass
import shutil
import requests
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

BASE_URL = "https://errolwatson.simprosuite.com"
COMPANY_ID = 0
TOKEN_URL = f"{BASE_URL}/oauth2/token"

CLIENT_ID = "260fa8bfed7a2aede3f4ae4dd4500b"
DEFAULT_QUOTE_TYPE = "Project"

COST_CENTRE_MAP = {
    "Plumbing Works": 6,
    "General Plumbing": 6,
    "Bathrooms": 4,
    "Heating": 5,
    "Joinery": 7,
    "Electrical": 11,
    "Painting": 13,
    "Painting & Decorating": 13,
    "Tiling": 14,
}

TASK_LABOUR_MAP = {
    "OUTSIDE-TAP": 104,
    "PIPE-CAP": 104,
    "ADD-HC-PIPE-M": 104,
    "PB-BATH-PANEL-700": 70,
    "WALL-CUT-M2": 70,
}

LABOR_TYPE_MAP = {
    "Plumbing": 104,
    "General Plumbing": 104,
    "Bathrooms": 104,
    "Heating": 104,
    "Joinery": 70,
    "Apprentice": 103,
}



def read_settings(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Settings"]

    return {
        "client_id": str(ws["B1"].value).strip(),
        "client_secret": str(ws["B2"].value).strip(),
        "company_id": int(ws["B3"].value),

        "auth_client_id": str(ws["B4"].value).strip(),
        "auth_client_secret": str(ws["B5"].value).strip(),
        "refresh_token": (Path.home() / "Downloads" / "simpro_refresh_token.txt").read_text().strip(),
        "refresh_token_path": str(Path.home() / "Downloads" / "simpro_refresh_token.txt"),
        "workbook_path": path
    }


def get_access_token(settings):
    global COMPANY_ID

    COMPANY_ID = settings["company_id"]

    response = requests.post(
        TOKEN_URL,
        data={
            "grant_type": "refresh_token",
            "client_id": settings["auth_client_id"],
            "client_secret": settings["auth_client_secret"],
            "refresh_token": settings["refresh_token"]
        }
    )

    print("Token status:", response.status_code)
    response.raise_for_status()

    token_data = response.json()

    new_refresh_token = token_data.get("refresh_token")
    if new_refresh_token and new_refresh_token != settings.get("refresh_token"):
        Path(settings["refresh_token_path"]).write_text(new_refresh_token)
        print("Refresh token updated in token file")

    return token_data["access_token"]


def headers(token):
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }


def clean(value):
    return str(value or "").strip()


def normalise_name(value):
    return clean(value).lower().replace(",", "")


def read_material_prices(wb):
    ws = wb["Materials List"]
    prices = {}

    for row in ws.iter_rows(min_row=11, values_only=True):
        code = clean(row[0])
        if not code:
            continue

        prices[code] = {
            "item": clean(row[3]),
            "spec": clean(row[4]),
            "unit": clean(row[5]),
            "trade_cost": row[6] or 0,
            "markup": row[7] or 0,
            "sell_ex_vat": row[8] or 0,
        }

    return prices


def read_task_materials(wb):
    ws = wb["Task Materials"]
    task_materials = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        task_code = clean(row[0])
        material_code = clean(row[1])

        if not task_code or not material_code:
            continue

        task_materials.setdefault(task_code, []).append({
            "material_code": material_code,
            "material_item": clean(row[2]),
            "fixed_qty": row[3] or 0,
            "qty_per_unit": row[4] or 0,
            "unit": clean(row[5]),
            "unit_sell_ex_vat": row[6] or 0,
            "notes": clean(row[7])
        })

    return task_materials


def read_final_quote(path):
    wb = load_workbook(path, data_only=True)

    final_ws = wb["Final Quote "]
    works_ws = wb["Works To Be Carried Out"]

    quote = {
        # V3 workbook layout
        "customer_name": clean(final_ws["B3"].value),
        "site_name": clean(final_ws["B4"].value),
        "lead_id": clean(final_ws["B5"].value),
        "description": clean(final_ws["B6"].value),
        "section_name": clean(final_ws["B8"].value) or "Works"
    }

    material_prices = read_material_prices(wb)
    task_materials = read_task_materials(wb)

    lines = []

    for row in works_ws.iter_rows(min_row=11, values_only=True):
        line_no = row[0]
        task_name = clean(row[2])
        task_code = clean(row[3])

        if not line_no or not task_name or not task_code:
            continue

        lines.append({
            "task_name": task_name,
            "task_code": task_code,
            "cost_centre_name": clean(row[1]),
            "count": row[4] or 1,
            "measurement": row[5] or 0,
            "unit": clean(row[6]),
            "labour_hours": row[8] or 0,
            "materials": task_materials.get(task_code, []),
            "material_prices": material_prices
        })

    return quote, lines


def get_customers(token):
    all_customers = []
    page = 1

    while True:
        response = requests.get(
            f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/customers/?pageSize=250&page={page}",
            headers=headers(token)
        )

        response.raise_for_status()
        customers = response.json()

        if not customers:
            break

        all_customers.extend(customers)

        if len(customers) < 250:
            break

        page += 1

    print("Customers checked:", len(all_customers))
    return all_customers


def find_customer(token, customer_name):
    customers = get_customers(token)

    target = normalise_name(customer_name)
    target_parts = sorted(target.split())

    for customer in customers:
        company_name = normalise_name(customer.get("CompanyName"))
        given = normalise_name(customer.get("GivenName"))
        family = normalise_name(customer.get("FamilyName"))

        full_name = f"{given} {family}".strip()
        reverse_name = f"{family} {given}".strip()

        if target == company_name:
            return customer

        if target == full_name:
            return customer

        if target == reverse_name:
            return customer

        if sorted(company_name.split()) == target_parts:
            return customer

        if sorted(full_name.split()) == target_parts:
            return customer

    return None


def looks_like_company(name):
    company_words = [
        "ltd", "limited", "group", "council", "company",
        "llp", "plc", "services", "homes", "estate", "estates"
    ]
    lower = name.lower()
    return any(word in lower for word in company_words)


def create_customer(token, customer_name):
    if looks_like_company(customer_name):
        url = f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/customers/companies/"
        payload = {"CompanyName": customer_name}
    else:
        parts = customer_name.split()
        given = parts[0] if parts else customer_name
        family = " ".join(parts[1:]) if len(parts) > 1 else ""

        url = f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/customers/individuals/"
        payload = {
            "GivenName": given,
            "FamilyName": family
        }

    response = requests.post(url, headers=headers(token), json=payload)
    print("Create customer status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()


def create_site(token, site_name):
    payload = {
        "Name": site_name,
        "Address": {
            "Address": site_name,
            "Country": "United Kingdom"
        }
    }

    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/sites/",
        headers=headers(token),
        json=payload
    )

    print("Create site status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()



def get_sites(token):
    all_sites = []
    page = 1

    while True:
        response = requests.get(
            f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/sites/?pageSize=250&page={page}",
            headers=headers(token)
        )

        response.raise_for_status()
        sites = response.json()

        if not sites:
            break

        all_sites.extend(sites)

        if len(sites) < 250:
            break

        page += 1

    print("Sites checked:", len(all_sites))
    return all_sites


def find_site(token, site_name):
    sites = get_sites(token)
    target = clean(site_name).lower()

    for site in sites:
        name = clean(site.get("Name")).lower()
        address = clean(site.get("Address", {}).get("Address")).lower()

        if target == name or target == address:
            return site

    return None


def create_quote(token, quote, customer_id, site_id):
    payload = {
        "Customer": customer_id,
        "Site": site_id,
        "Type": DEFAULT_QUOTE_TYPE,
        "Name": quote["description"],
        "Description": quote["description"]
    }

    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/quotes/",
        headers=headers(token),
        json=payload
    )

    print("Quote create status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()


def create_section(token, quote_id, section_name):
    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/quotes/{quote_id}/sections/",
        headers=headers(token),
        json={"Name": section_name}
    )

    print("Section status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()


def add_cost_centre(token, quote_id, section_id, cost_centre_id):
    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/quotes/{quote_id}/sections/{section_id}/costCenters/",
        headers=headers(token),
        json={"CostCenter": cost_centre_id}
    )

    print("Cost centre status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()


def add_labor(token, quote_id, section_id, quote_cost_centre_id, labor_type_id, hours):
    try:
        hours = float(hours or 0)
    except (TypeError, ValueError):
        return

    if hours <= 0:
        return None

    payload = {
        "LaborType": labor_type_id,
        "Total": {
            "Qty": hours
        }
    }

    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/quotes/{quote_id}/sections/{section_id}/costCenters/{quote_cost_centre_id}/labor/",
        headers=headers(token),
        json=payload
    )

    print("Labour status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()


def add_material(token, quote_id, section_id, quote_cost_centre_id, description, qty, sell_price):
    if not qty or qty <= 0:
        return None

    payload = {
        "Type": "Material",
        "Description": description,
        "Total": {
            "Qty": qty
        },
        "SellPrice": sell_price or 0
    }

    response = requests.post(
        f"{BASE_URL}/api/v1.0/companies/{COMPANY_ID}/quotes/{quote_id}/sections/{section_id}/costCenters/{quote_cost_centre_id}/oneOffs/",
        headers=headers(token),
        json=payload
    )

    print("Material status:", response.status_code)
    print(response.text)
    response.raise_for_status()
    return response.json()



def safe_filename(value):
    bad_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    value = clean(value)
    for ch in bad_chars:
        value = value.replace(ch, "-")
    return value[:80]


def archive_and_reset_workbook(workbook_path, quote_id, quote):
    """Archive a copy only. Do not edit/reset the live workbook.

    This avoids using openpyxl to save the macro workbook, which can damage
    dropdowns/data validation. Once the send flow is stable, reset can be added
    safely via Excel/VBA instead.
    """
    workbook = Path(workbook_path)
    clients_root = Path("/Users/ewgcoomercial/Library/CloudStorage/OneDrive-ErrolWatsonGroup/Clients")

    customer_folder_name = safe_filename(quote.get("customer_name", "Customer"))
    first_letter = customer_folder_name[:1].upper() if customer_folder_name else "Unknown"

    archive_folder = clients_root / first_letter / customer_folder_name
    archive_folder.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    customer = safe_filename(quote.get("customer_name", "Customer"))
    site = safe_filename(quote.get("site_name", "Site"))

    archive_name = f"{quote_id} - {customer} - {site} - {today}.xlsm"
    archive_path = archive_folder / archive_name

    shutil.copy2(workbook_path, archive_path)

    print("")
    print("SUCCESS")
    print(f"Quote {quote_id} created in simPRO")
    print(f"Workbook archived to: {archive_path}")
    print("Live workbook was NOT reset or edited.")


def get_labor_type_for_line(line):
    if line["task_code"] in TASK_LABOUR_MAP:
        return TASK_LABOUR_MAP[line["task_code"]]

    return LABOR_TYPE_MAP.get(line["cost_centre_name"], 104)


def main():
    import sys

    workbook_path = sys.argv[1]
    quote, lines = read_final_quote(workbook_path)

    print("Customer:", quote["customer_name"])
    print("Site:", quote["site_name"])
    print("Lines found:", len(lines))

    settings = read_settings(workbook_path)
    token = get_access_token(settings)

    customer = find_customer(token, quote["customer_name"])

    if customer:
        print("Found existing customer:", customer["ID"], customer.get("CompanyName", ""))
    else:
        print("Customer not found. Creating new customer.")
        customer = create_customer(token, quote["customer_name"])

    site = find_site(token, quote["site_name"])

    if site:
        print("Found existing site:", site["ID"], site.get("Name", ""))
    else:
        print("Site not found. Creating new site.")
        site = create_site(token, quote["site_name"])

    created_quote = create_quote(token, quote, customer["ID"], site["ID"])
    quote_id = created_quote["ID"]

    print("Created quote:", quote_id)

    section = create_section(token, quote_id, quote["section_name"])
    section_id = section["ID"]

    created_cost_centres = {}

    for line in lines:
        cost_centre_id = COST_CENTRE_MAP.get(line["cost_centre_name"], 6)

        if cost_centre_id not in created_cost_centres:
            quote_cost_centre = add_cost_centre(
                token,
                quote_id,
                section_id,
                cost_centre_id
            )
            created_cost_centres[cost_centre_id] = quote_cost_centre["ID"]

        quote_cost_centre_id = created_cost_centres[cost_centre_id]
        labor_type_id = get_labor_type_for_line(line)

        print(
            "Adding labour:",
            line["task_name"],
            line["labour_hours"],
            "hrs",
            "LaborType:",
            labor_type_id
        )

        add_labor(
            token,
            quote_id,
            section_id,
            quote_cost_centre_id,
            labor_type_id,
            line["labour_hours"]
        )

        for material in line["materials"]:
            qty = material["fixed_qty"] + (material["qty_per_unit"] * line["measurement"])
            qty = qty * line["count"]

            material_code = material["material_code"]
            price_info = line["material_prices"].get(material_code, {})

            sell_price = material["unit_sell_ex_vat"] or price_info.get("sell_ex_vat", 0) or 0

            material_name = price_info.get("item") or material["material_item"]
            spec = price_info.get("spec", "")
            unit = material["unit"] or price_info.get("unit", "")

            description = f"{material_name} {spec}".strip()
            description = f"{description} | Code: {material_code} | Unit: {unit}"

            if sell_price == 0:
                description += " | PRICE MISSING"

            print("Adding material:", description, "Qty:", qty, "Sell:", sell_price)

            add_material(
                token,
                quote_id,
                section_id,
                quote_cost_centre_id,
                description,
                qty,
                sell_price
            )

    archive_and_reset_workbook(workbook_path, quote_id, quote)

    print("Finished creating quote with native labour and itemised materials.")


if __name__ == "__main__":
    main()
